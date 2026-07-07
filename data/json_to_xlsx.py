#!/usr/bin/env python3
"""
Convert an app trip JSON (stays/travels/days model) into a realistic, flat
itinerary spreadsheet (.xlsx) that resembles something a user would upload.

The goal is a TEST FIXTURE for the AI import pipeline, so it deliberately:
  - drops internal UUIDs (dayId/pointId/locationId)
  - drops enrichment fields (lat/lng, googlePlaceId, googleMapsUri, fullAddress)
  - drops long descriptions (the AI "enhance" pass is expected to fill these in)

It keeps the factual scaffolding a traveller would type: dates, titles, times,
point action type, travel mode / stay type, confirmation numbers and location names.

Usage:
    python3 data/json_to_xlsx.py [input.json] [output.xlsx]

Defaults: data/trip.json -> data/trip_example.xlsx
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "Date",
    "Day",
    "Point Type",
    "Title",
    "Start Time",
    "End Time",
    "Travel Mode",
    "Stay Type",
    "Confirmation #",
    "Locations",
]


def _time_only(dt: str | None) -> str:
    """Return HH:MM from an ISO datetime, or '' if missing."""
    if not dt:
        return ""
    # e.g. "2026-05-11T12:15:00+02:00" -> "12:15"
    try:
        return dt.split("T", 1)[1][:5]
    except IndexError:
        return ""


def _locations(locations: list[dict]) -> str:
    parts = []
    for loc in locations:
        name = loc.get("name", "").strip()
        if not name:
            continue
        role = loc.get("role")
        parts.append(f"{name} ({role})" if role else name)
    return "; ".join(parts)


def build_rows(trip: dict) -> list[list[str]]:
    stays = {s.get("stayDetailId"): s for s in trip.get("stays", []) if s.get("stayDetailId")}
    travels = {t.get("travelDetailId"): t for t in trip.get("travels", []) if t.get("travelDetailId")}

    rows: list[list[str]] = []
    for day in trip.get("days", []):
        day_title = day.get("title", "")
        date = day.get("date", "")
        points = day.get("points", [])
        if not points:
            rows.append([date, day_title, "", "", "", "", "", "", "", ""])
            continue
        for point in points:
            travel = travels.get(point.get("travelDetailId")) or {}
            stay = stays.get(point.get("stayDetailId")) or {}
            merged_locations = [
                *(point.get("locations") or []),
                *(travel.get("locations") or []),
                *(stay.get("locations") or []),
            ]
            confirmation = (
                point.get("confirmationNumber")
                or travel.get("confirmationNumber")
                or stay.get("confirmationNumber")
                or ""
            )
            rows.append(
                [
                    date,
                    day_title,
                    point.get("type", ""),
                    point.get("title", ""),
                    _time_only(point.get("startDateTime")),
                    _time_only(point.get("endDateTime")),
                    travel.get("mode", "") or "",
                    stay.get("stayType", "") or "",
                    confirmation,
                    _locations(merged_locations),
                ]
            )
    return rows


def write_workbook(trip: dict, rows: list[list[str]], out_path: Path) -> None:
    wb = Workbook()

    # ── Trip summary sheet ────────────────────────────────────────────────
    summary = wb.active
    summary.title = "Trip"
    summary["A1"] = "Trip Name"
    summary["B1"] = trip.get("tripName", "")
    summary["A2"] = "Start Date"
    summary["B2"] = trip.get("startDate", "")
    summary["A3"] = "End Date"
    summary["B3"] = trip.get("endDate", "")
    for r in range(1, 4):
        summary[f"A{r}"].font = Font(bold=True)
    summary.column_dimensions["A"].width = 16
    summary.column_dimensions["B"].width = 48

    # ── Itinerary sheet (flat, user-style) ────────────────────────────────
    ws = wb.create_sheet("Itinerary")
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    widths = [12, 26, 10, 40, 11, 11, 13, 12, 18, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cells in ws.iter_rows(min_row=1):
        for cell in cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"

    wb.save(out_path)


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Convert a trip JSON into an itinerary spreadsheet.")
    parser.add_argument("input", nargs="?", default="data/trip.json", help="Input trip JSON path.")
    parser.add_argument("output", nargs="?", default="data/trip_example.xlsx", help="Output .xlsx path.")
    parser.add_argument(
        "--gap-day",
        type=int,
        default=None,
        metavar="N",
        help="Empty the legs of the Nth day (1-based) to create a fully-empty day for testing verification.",
    )
    args = parser.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)

    trip = json.loads(in_path.read_text(encoding="utf-8"))

    if args.gap_day is not None:
        days = trip.get("days", [])
        idx = args.gap_day - 1
        if not 0 <= idx < len(days):
            parser.error(f"--gap-day {args.gap_day} is out of range (1..{len(days)}).")
        days[idx]["points"] = []
        print(f"Emptied day {args.gap_day}: {days[idx].get('title', '')}")

    rows = build_rows(trip)
    write_workbook(trip, rows, out_path)
    print(f"Wrote {out_path} ({len(rows)} itinerary rows across {len(trip.get('days', []))} days)")


if __name__ == "__main__":
    main()
